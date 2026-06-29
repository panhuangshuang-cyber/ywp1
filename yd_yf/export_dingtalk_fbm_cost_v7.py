#!/usr/bin/env python3
"""
钉钉「单SKU发货成本」导出 v7（可 import 版）
新增 get_fbm_cost_df() 接口，方便被运费核查模块 import
"""

import requests
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime
import urllib.parse
import time

# ==================== 配置（同 v6）====================
WORKBOOK_ID = "vy20BglGWOgznx2Ms0ENN6R5WA7depqY"
SHEET_NAME = "单SKU发货成本"
OPERATOR_ID = "o6cNzAqelYyzTaxc2KcIfgiEiE"

APPKEY = "dinglhtnqselknfytfqq"
APPSECRET = "S_qytG7eF3SylcE5nRzYTbSMJnayQbJ07nicye3nRhFzzsfbHWGbcCkZNFl0DYKG"

BATCH_SIZE = 300
OUTPUT_FILE = f"单SKU发货成本_最终版_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
# =============================================


def get_access_token():
    url = f"https://oapi.dingtalk.com/gettoken?appkey={APPKEY}&appsecret={APPSECRET}"
    resp = requests.get(url, timeout=10).json()
    if "access_token" not in resp:
        raise Exception(f"Token失败: {resp}")
    return resp["access_token"]


def forward_fill_row(row):
    filled = []
    prev = ""
    for val in row:
        val = str(val).strip() if val is not None else ""
        if not val:
            val = prev
        filled.append(val)
        if val:
            prev = val
    return filled


def build_smart_header(row8, row9):
    row8_filled = forward_fill_row(row8)
    row9_filled = forward_fill_row(row9)
    headers = []
    prev = ""
    for i in range(max(len(row8_filled), len(row9_filled))):
        h1 = row8_filled[i] if i < len(row8_filled) else ""
        h2 = row9_filled[i] if i < len(row9_filled) else ""
        if h1 and h2 and h1 != h2:
            combined = f"{h1} / {h2}"
        else:
            combined = h1 or h2 or prev
        if not combined:
            combined = prev
        headers.append(combined)
        prev = combined
    return headers


def clean_column_name(name: str, existing: set) -> str:
    if not name or not str(name).strip():
        base = "col_unknown"
    else:
        s = str(name).strip()
        s = re.sub(r'[\s/\\\-–—|:,;()（）【】\[\]""''「」『』\.\*\+\?]+', '_', s)
        s = re.sub(r'_+', '_', s)
        s = s.strip('_')
        base = s if s else "col_unknown"
    final = base
    suffix = 1
    while final in existing:
        final = f"{base}_{suffix}"
        suffix += 1
    existing.add(final)
    return final


def fetch_range(token, sheet_name_encoded, start_row, end_row):
    headers = {
        "content-type": "application/json",
        "x-acs-dingtalk-access-token": token
    }
    range_url = (
        f"https://api.dingtalk.com/v1.0/doc/workbooks/{WORKBOOK_ID}/sheets/{sheet_name_encoded}/ranges/"
        f"A{start_row}:AD{end_row}?operatorId={OPERATOR_ID}&select=values"
    )
    resp = requests.get(range_url, headers=headers, timeout=30)
    data_json = resp.json()
    if "errCode" in data_json and data_json.get("errCode") != 200:
        print(f"【范围 {start_row}-{end_row} 错误】errCode={data_json.get('errCode')} msg={data_json.get('errMsg')}")
        return []
    return data_json.get("values", [])


def fetch_and_process():
    token = get_access_token()
    sheet_name_encoded = urllib.parse.quote(SHEET_NAME)

    headers = {
        "content-type": "application/json",
        "x-acs-dingtalk-access-token": token
    }
    row_url = f"https://api.dingtalk.com/v1.0/doc/workbooks/{WORKBOOK_ID}/sheets/{sheet_name_encoded}?operatorId={OPERATOR_ID}"
    row_resp = requests.get(row_url, headers=headers, timeout=15).json()
    total_row = row_resp.get("rowCount", 0)
    print(f"总行数: {total_row}")

    if total_row < 11:
        print("数据不足")
        return pd.DataFrame(), []

    header_values = fetch_range(token, sheet_name_encoded, 1, 15)
    if not header_values or len(header_values) < 10:
        print("表头区域拉取失败")
        return pd.DataFrame(), []

    header_row8 = header_values[8]
    header_row9 = header_values[9]
    raw_headers = build_smart_header(header_row8, header_row9)

    existing_names = set()
    clean_headers = [clean_column_name(h, existing_names) for h in raw_headers]
    print(f"清洗后表头示例: {clean_headers[:6]}")

    all_data_values = []
    start = 11
    while start <= total_row:
        end = min(start + BATCH_SIZE - 1, total_row)
        batch = fetch_range(token, sheet_name_encoded, start, end)
        if batch:
            all_data_values.extend(batch)
        print(f"已拉取 {start}-{end} 行，累计 {len(all_data_values)} 行")
        start = end + 1
        time.sleep(0.3)

    print(f"实际数据行: {len(all_data_values)}")

    if not all_data_values:
        return pd.DataFrame(), clean_headers

    records = []
    for row in all_data_values:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        record = {}
        for i, h in enumerate(clean_headers):
            val = row[i] if i < len(row) else ""
            record[h] = val
        records.append(record)

    df = pd.DataFrame(records)

    sku_col = next((c for c in clean_headers if "sku" in c.lower() or "型号" in c), None)
    if sku_col:
        df = df[df[sku_col].astype(str).str.strip() != ""]

    print(f"有效数据行: {len(df)}")
    return df, clean_headers


def get_fbm_cost_df() -> pd.DataFrame:
    """
    可被其他模块 import 的接口
    返回标准 DataFrame，列名已清洗，包含 SKU 和 FBM_尾程运费 等
    """
    df, clean_headers = fetch_and_process()
    if df.empty:
        return df

    # 优先使用标准列名，避免误选「供应商型号」等字段
    sku_col = next((c for c in df.columns if str(c).strip().upper() == "SKU"), None)
    if not sku_col:
        sku_col = next((c for c in df.columns if "sku" in str(c).lower()), None)

    fbm_col = next((c for c in df.columns if str(c) == "FBM_尾程运费"), None)
    if not fbm_col:
        fbm_col = next(
            (c for c in df.columns if "fbm" in str(c).lower() and "尾程" in str(c)),
            None,
        )

    if sku_col and fbm_col:
        result = df[[sku_col, fbm_col]].rename(columns={sku_col: "SKU", fbm_col: "FBM_尾程运费"})
        result['SKU'] = result['SKU'].astype(str).str.strip()
        result['FBM_尾程运费'] = pd.to_numeric(result['FBM_尾程运费'], errors='coerce')
        result = result[result['SKU'].ne('') & result['SKU'].ne('nan')]
        return result.drop_duplicates(subset=['SKU'], keep='first').reset_index(drop=True)
    return df


def save_professional_excel(df: pd.DataFrame, headers: list, filename: str):
    if df.empty:
        print("无有效数据，跳过生成 Excel")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "单SKU发货成本"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center")

    for col_idx in range(1, ws.max_column + 1):
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = '#,##0.00'

    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for col_idx, h in enumerate(headers, 1):
        h_lower = h.lower()
        if any(kw in h_lower for kw in ["fbm", "运费", "尾程", "shipping", "fee"]):
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).fill = highlight_fill
            print(f"已高亮: {h}")

    ws.freeze_panes = "A2"
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=min(ws.max_column, 35)):
        for cell in row_cells:
            cell.border = thin

    for col_idx in range(1, min(ws.max_column + 1, 35)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(filename)
    print(f"✅ 已导出: {filename}（共 {len(df)} 行）")


if __name__ == "__main__":
    print("=== 单SKU发货成本 v7（可 import 版）开始 ===")
    df = get_fbm_cost_df()
    if not df.empty:
        # 同时保存一份 xlsx 供手动查看
        _, headers = fetch_and_process()
        save_professional_excel(df, headers, OUTPUT_FILE)
    print("=== 完成 ===")
    print("其他模块可 from export_dingtalk_fbm_cost_v7 import get_fbm_cost_df")