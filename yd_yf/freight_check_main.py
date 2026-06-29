#!/usr/bin/env python3
"""
运费核查主流程
串联：运德订单 + SKU 别名 + 钉钉成本表 + SKU 白名单 -> 差异报告
"""

from __future__ import annotations

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from export_dingtalk_fbm_cost_v7 import get_fbm_cost_df
from sku_alias import get_alias_map, normalize_sku_list
from yd_order_api import (
    build_merge_sku_key,
    clean_orders,
    fetch_wedo_orders,
    get_order_date_range,
    parse_goods_info,
)

WEDO_EXPORT_HEADERS = [
    '出库单号', '生成时间', '运费', '数量',
    '产品编号', '数量', '产品编号', '数量', '产品编号',
]

BASE_DIR = os.path.dirname(__file__)
SKU_LIST_PATH = os.path.join(BASE_DIR, "需核查运费SKU.xlsx")
REPORT_DIR = os.path.join(BASE_DIR, "artifacts")
DIFF_THRESHOLD = 2


def period_column_names(date_label: str) -> tuple[str, str]:
    count_col = f"过去90天出单总数\n（{date_label}）"
    freight_col = f"过去90天出单总运费\n（{date_label}）"
    return count_col, freight_col


def report_column_names(date_label: str) -> list[str]:
    count_col, freight_col = period_column_names(date_label)
    return [
        'SKU', '别名', count_col, freight_col,
        '平均运费', '产品成本表运费', '差异', '合并运费计算款说明',
    ]


def load_sku_whitelist(path: str = SKU_LIST_PATH) -> pd.DataFrame:
    df = pd.read_excel(path)
    if 'SKU' not in df.columns:
        raise ValueError(f"白名单缺少 SKU 列: {path}")

    result = df.copy()
    result['SKU'] = result['SKU'].astype(str).str.strip()
    result = result[result['SKU'].ne('') & result['SKU'].ne('nan')]
    if '合并运费计算款说明' in result.columns:
        result['可合并运费'] = result['合并运费计算款说明'].astype(str).str.contains('可合并运费', na=False)
    else:
        result['可合并运费'] = False
    return result.reset_index(drop=True)


def orders_to_detail(df_orders: pd.DataFrame, alias_map: dict[str, str]) -> pd.DataFrame:
    if df_orders.empty:
        return pd.DataFrame(columns=['order_no', 'SKU', 'freight', 'create_time'])

    order_col = 'orderId' if 'orderId' in df_orders.columns else 'order_no'
    fee_col = 'shipfee' if 'shipfee' in df_orders.columns else 'freight'
    time_col = 'createTime' if 'createTime' in df_orders.columns else 'create_time'

    records: list[dict] = []
    for _, row in df_orders.iterrows():
        raw_skus = row.get('user_skus') or []
        if not raw_skus:
            continue
        skus = normalize_sku_list(list(raw_skus), alias_map)
        if not skus:
            continue
        sku_key = build_merge_sku_key(skus)
        if not sku_key:
            continue
        records.append({
            'order_no': str(row.get(order_col, '')),
            'SKU': sku_key,
            'freight': float(row.get(fee_col, 0) or 0),
            'create_time': row.get(time_col, ''),
        })

    return pd.DataFrame(records)


def aggregate_orders(df_detail: pd.DataFrame) -> pd.DataFrame:
    if df_detail.empty:
        return pd.DataFrame(columns=['SKU', 'order_count', 'freight_sum', 'avg_freight'])

    grouped = df_detail.groupby('SKU', as_index=False).agg(
        order_count=('order_no', 'count'),
        freight_sum=('freight', 'sum'),
    )
    grouped['avg_freight'] = grouped['freight_sum'] / grouped['order_count']
    return grouped


def _sku_components(sku_key: str) -> list[str]:
    return [part.strip() for part in str(sku_key).split('/') if part.strip()]


def _match_whitelist(df_agg: pd.DataFrame, df_whitelist: pd.DataFrame) -> pd.DataFrame:
    whitelist_skus = set(df_whitelist['SKU'].tolist())
    merge_skus = set(df_whitelist.loc[df_whitelist['可合并运费'], 'SKU'].tolist())

    def keep_row(sku_key: str) -> bool:
        if sku_key in whitelist_skus:
            return True
        parts = _sku_components(sku_key)
        if len(parts) <= 1:
            return False
        if any(part in merge_skus for part in parts):
            return True
        return any(part in whitelist_skus for part in parts)

    if df_agg.empty:
        return df_agg
    mask = df_agg['SKU'].map(keep_row)
    return df_agg[mask].reset_index(drop=True)


def lookup_fbm_cost(sku_key: str, cost_lookup: dict[str, float]) -> float | None:
    if sku_key in cost_lookup:
        return cost_lookup[sku_key]

    parts = _sku_components(sku_key)
    if len(parts) <= 1:
        return cost_lookup.get(sku_key)

    values = [cost_lookup[part] for part in parts if part in cost_lookup]
    return float(sum(values)) if values else None


def calculate_diff(df_agg: pd.DataFrame, df_cost: pd.DataFrame) -> pd.DataFrame:
    cost_lookup: dict[str, float] = {}
    if not df_cost.empty and {'SKU', 'FBM_尾程运费'}.issubset(df_cost.columns):
        for _, row in df_cost.iterrows():
            sku = str(row['SKU']).strip()
            fbm = pd.to_numeric(row['FBM_尾程运费'], errors='coerce')
            if sku and pd.notna(fbm):
                cost_lookup[sku] = float(fbm)

    rows = []
    for _, row in df_agg.iterrows():
        sku_key = str(row['SKU'])
        product_cost = lookup_fbm_cost(sku_key, cost_lookup)
        avg_freight = float(row['avg_freight']) if pd.notna(row['avg_freight']) else None
        diff = (avg_freight - product_cost) if avg_freight is not None and product_cost is not None else None
        rows.append({
            'SKU': sku_key,
            'order_count': row['order_count'],
            'freight_sum': row['freight_sum'],
            'avg_freight': avg_freight,
            'product_cost': product_cost,
            'diff': diff,
            'highlight': abs(diff) > DIFF_THRESHOLD if diff is not None else False,
        })
    return pd.DataFrame(rows)


def build_report_dataframe(
    df_whitelist: pd.DataFrame,
    df_stats: pd.DataFrame,
    date_label: str,
) -> pd.DataFrame:
    """按需核查运费SKU.xlsx 列结构生成报告"""
    count_col, freight_col = period_column_names(date_label)
    report_cols = report_column_names(date_label)

    base = df_whitelist[['SKU', '别名', '合并运费计算款说明']].copy()
    for col in [count_col, freight_col, '平均运费', '产品成本表运费', '差异']:
        base[col] = None

    if not df_stats.empty:
        stats = df_stats.set_index('SKU')
        for idx, row in base.iterrows():
            sku = str(row['SKU'])
            if sku not in stats.index:
                continue
            stat = stats.loc[sku]
            base.at[idx, count_col] = stat['order_count']
            base.at[idx, freight_col] = stat['freight_sum']
            base.at[idx, '平均运费'] = stat['avg_freight']
            base.at[idx, '产品成本表运费'] = stat['product_cost']
            base.at[idx, '差异'] = stat['diff']

    whitelist_set = set(base['SKU'].astype(str))
    composite_stats = df_stats[~df_stats['SKU'].astype(str).isin(whitelist_set)] if not df_stats.empty else pd.DataFrame()
    extra_rows: list[dict] = []
    for _, stat in composite_stats.iterrows():
        extra_rows.append({
            'SKU': stat['SKU'],
            '别名': None,
            count_col: stat['order_count'],
            freight_col: stat['freight_sum'],
            '平均运费': stat['avg_freight'],
            '产品成本表运费': stat['product_cost'],
            '差异': stat['diff'],
            '合并运费计算款说明': None,
        })

    report = pd.concat([base, pd.DataFrame(extra_rows)], ignore_index=True)
    return report[report_cols]


def _parse_order_time(value) -> datetime | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        ts = int(value)
        if ts > 1_000_000_000_000:
            ts //= 1000
        return datetime.fromtimestamp(ts)

    text = str(value).strip()
    if not text or text.lower() == 'nan':
        return None

    numeric = pd.to_numeric(text, errors='coerce')
    if pd.notna(numeric):
        ts = int(numeric)
        if ts > 1_000_000_000_000:
            ts //= 1000
        return datetime.fromtimestamp(ts)

    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def _format_order_time(value) -> str:
    """导出用时间文本，避免 Excel 显示 unix 时间戳数字"""
    dt = _parse_order_time(value)
    return dt.strftime('%Y-%m-%d %H:%M:%S') if dt else ''


def _extract_order_products(row: pd.Series) -> list[tuple[str, int]]:
    goods_col = 'goodsInfo' if 'goodsInfo' in row.index else 'goods_info'
    products: list[tuple[str, int]] = []
    for item in parse_goods_info(row.get(goods_col)):
        sku = str(item.get('userSku') or '').strip()
        if not sku:
            continue
        qty = pd.to_numeric(item.get('qty', 1), errors='coerce')
        products.append((sku, int(qty) if pd.notna(qty) else 1))
    return products


def _order_to_export_row(row: pd.Series) -> list:
    """
    按运德导出模板展开一行：
    单品：产品编号 + 空数量列；多品：前两列带数量，第三个产品编号不带数量列。
    """
    products = _extract_order_products(row)
    if len(products) > 3:
        order_no = row.get('orderId', row.get('order_no', ''))
        print(f"[警告] 订单 {order_no} 含 {len(products)} 个产品，模板仅支持 3 列，已截断")

    order_id = row.get('orderId', row.get('order_no', ''))
    create_time_raw = row.get('createTime', row.get('create_time'))
    create_time_text = _format_order_time(create_time_raw)
    freight = pd.to_numeric(row.get('shipfee', row.get('freight')), errors='coerce')

    export_row: list = [order_id, create_time_text, freight, 1]
    product_cells: list = [None, None, None, None, None]

    if len(products) == 1:
        product_cells = [products[0][0], None, None, None, None]
    elif len(products) == 2:
        product_cells = [products[0][0], products[0][1], products[1][0], products[1][1], None]
    elif len(products) >= 3:
        product_cells = [
            products[0][0], products[0][1],
            products[1][0], products[1][1],
            products[2][0],
        ]

    return export_row + product_cells


def export_clean_orders_excel(
    df_orders: pd.DataFrame,
    date_label: str,
    output_dir: str = REPORT_DIR,
) -> str:
    """按运德导出模板格式输出清洗后的 90 天订单"""
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"运德90天_{date_label}.xlsx")

    indexed_rows: list[tuple[datetime, list]] = []
    for _, row in df_orders.iterrows():
        sort_time = _parse_order_time(row.get('createTime', row.get('create_time'))) or datetime.min
        indexed_rows.append((sort_time, _order_to_export_row(row)))
    indexed_rows.sort(key=lambda item: item[0], reverse=True)
    export_rows = [values for _, values in indexed_rows]

    wb = Workbook()
    ws = wb.active
    ws.title = "运德90天"

    for col_idx, header in enumerate(WEDO_EXPORT_HEADERS, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, values in enumerate(export_rows, 2):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✅ 运德清洗订单已导出: {output_path}（{len(export_rows)} 条）")
    return output_path


def generate_report(df_result: pd.DataFrame, date_label: str, output_path: str | None = None) -> str:
    os.makedirs(REPORT_DIR, exist_ok=True)
    if output_path is None:
        output_path = os.path.join(
            REPORT_DIR,
            f"运费核查报告_{date_label}_{datetime.now().strftime('%H%M')}.xlsx",
        )

    report_cols = report_column_names(date_label)
    display_df = df_result.copy()
    for col in report_cols:
        if col not in display_df.columns:
            display_df[col] = None
    display_df = display_df[report_cols]

    highlight_mask = display_df['差异'].apply(
        lambda value: pd.notna(value) and abs(float(value)) > DIFF_THRESHOLD
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "运费核查结果"

    for r_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row_idx in range(2, ws.max_row + 1):
        if highlight_mask.iloc[row_idx - 2]:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"✅ 报告已生成: {output_path}")
    return output_path


def run_freight_check(days: int = 90) -> pd.DataFrame:
    print("=== 运费核查主流程 开始 ===")

    _, _, date_label = get_order_date_range(days)
    count_col, freight_col = period_column_names(date_label)
    print(f"[区间] {date_label}")

    df_whitelist = load_sku_whitelist()
    alias_map = get_alias_map(SKU_LIST_PATH)
    df_orders = clean_orders(fetch_wedo_orders(days=days))
    df_cost = get_fbm_cost_df()

    print(
        f"[数据] 白名单 {len(df_whitelist)} 个 | 订单 {len(df_orders)} 条 | "
        f"成本 {len(df_cost)} 条 | 别名 {len(alias_map)} 条"
    )

    export_clean_orders_excel(df_orders, date_label)

    df_detail = orders_to_detail(df_orders, alias_map)
    df_agg = aggregate_orders(df_detail)
    df_agg = _match_whitelist(df_agg, df_whitelist)
    df_stats = calculate_diff(df_agg, df_cost)
    df_report = build_report_dataframe(df_whitelist, df_stats, date_label)

    highlight_count = int(df_stats['highlight'].fillna(False).sum()) if not df_stats.empty else 0
    print(f"[结果] 聚合 SKU {len(df_agg)} 个 | 差异>{DIFF_THRESHOLD} 高亮 {highlight_count} 条")
    print(f"[列名] {count_col} / {freight_col}")

    generate_report(df_report, date_label)
    print("=== 完成 ===")
    return df_report


def main() -> pd.DataFrame:
    return run_freight_check(days=90)


if __name__ == "__main__":
    main()